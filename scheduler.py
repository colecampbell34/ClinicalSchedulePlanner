import pandas as pd
import random
from dataclasses import dataclass, field
from typing import List, Dict
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

"""
Clinical Placement Scheduler

This module automates the scheduling of clinical student placements into hospital sites 
and clinics. It handles complex constraints for different student streams (General, 
Cardiac, Dual), balances site capacities, optimizes for student choices, and generates 
a color-coded Excel schedule.
"""


# ==================== DATA CLASSES ====================

@dataclass
class Site:
    """
    Represents a hospital placement site with specific capabilities and constraints.

    Attributes:
        site_id (float): Unique identifier for the site.
        name (str): Name of the hospital/site.
        is_general (bool): Whether the site supports General stream placements.
        is_cardiac (bool): Whether the site supports Cardiac stream placements.
        is_dual (bool): Whether the site supports Dual stream (CS/Perf) placements.
        high_acuity (bool): Indicates if the site handles high-acuity cases.
        out_of_town (bool): Indicates if the site is outside the main metro area.
        only_may_to_dec (bool): If True, site is unavailable in Jan-April blocks (2A-3A).
        capacity (int): Maximum number of students the site can take per block.
    """
    site_id: float
    name: str
    is_general: bool
    is_cardiac: bool
    is_dual: bool
    high_acuity: bool
    out_of_town: bool
    only_may_to_dec: bool
    capacity: int


@dataclass
class Clinic:
    """
    Represents a smaller clinical rotation option.

    Attributes:
        clinic_id (float): Unique identifier for the clinic.
        name (str): Name of the clinic.
        days_per_week (int): Number of days per week the placement runs.
        out_of_town (bool): Indicates if the clinic is outside the main metro area.
        only_may_to_dec (bool): If True, clinic is unavailable in Jan-April blocks.
    """
    clinic_id: float
    name: str
    days_per_week: int
    out_of_town: bool
    only_may_to_dec: bool


@dataclass
class Student:
    """
    Represents a student requiring placement.

    Attributes:
        student_id (str): Unique student ID.
        last_name (str): Student's last name.
        first_name (str): Student's first name.
        program (str): Program stream ("General", "Cardiac", or "Dual").
        choices (List[str]): Ordered list of preferred site names (1st to 5th).
        accommodation_sites (List[str]): List of out-of-town sites where student has housing.
        cardiac_weeks_completed (int): Cumulative cardiac clinical weeks completed.
        clinic_blocks (int): Number of clinic blocks assigned.
        has_had_ob_exposure (bool): Tracks if student has completed a block with OB exposure.
        is_chain_starter (bool): Flag for Dual students authorized to start a multi-block chain.
        placements (Dict[str, str]): Map of Block ID to Site Name assignment.
        assigned_sites (List[str]): History of all assigned site names to prevent unwanted repetition.
    """
    student_id: str
    last_name: str
    first_name: str
    program: str
    choices: List[str]
    accommodation_sites: List[str]

    cardiac_weeks_completed: int = 0
    clinic_blocks: int = 0
    has_had_ob_exposure: bool = False

    # Track if this student is authorized for a Dual Chain
    is_chain_starter: bool = False

    placements: Dict[str, str] = field(default_factory=dict)
    assigned_sites: List[str] = field(default_factory=list)


# ==================== SCHEDULER ====================

class Scheduler:
    """
    The core engine for generating the schedule.

    This class loads data, manages state (site usage, student progress),
    runs the scheduling algorithms per block, and formats the output.
    """

    def __init__(self, excel_path: str):
        """
        Initialize the Scheduler.

        Args:
            excel_path (str): Path to the input Excel file containing
                              'Students', 'Sites', and 'Clinics' sheets.
        """
        self.excel_path = excel_path
        self.students = []
        self.sites = []
        self.clinics = []

        self.cardiac_names = set()
        self.clinic_names = set()
        self.general_names = set()
        self.dual_site_names = set()

        # (Student_ID, Block) -> "Red" | "Blue" | "Green" | "Orange" | "Purple" | "Warn"
        self.placement_intent = {}

        # Tracking
        self.site_slots = {b: defaultdict(list) for b in BLOCKS}
        self.clinic_days_used = {b: defaultdict(int) for b in BLOCKS}
        self.site_name_usage = {b: defaultdict(int) for b in BLOCKS}

        # QUOTA TRACKER
        self.dual_chains_assigned = 0
        self.MAX_DUAL_CHAINS = 5

        # ==================== LOAD ====================

    def load_data(self):
        """
        Reads student, site, and clinic data from the Excel file.
        Populates the internal lists and sets for rapid lookup.
        """
        try:
            students_df = pd.read_excel(self.excel_path, sheet_name="Students")
            sites_df = pd.read_excel(self.excel_path, sheet_name="Sites")
            clinics_df = pd.read_excel(self.excel_path, sheet_name="Clinics")
        except FileNotFoundError:
            print(f"Error: Could not find file '{self.excel_path}'")
            return

        for _, r in students_df.iterrows():
            choices = [
                str(r[f"choice_{i}"]).strip()
                for i in range(1, 6)
                if f"choice_{i}" in r and pd.notna(r[f"choice_{i}"])
            ]

            acc = []
            if pd.notna(r.get("accommodation_sites")):
                acc = [s.strip() for s in str(r["accommodation_sites"]).split(",")]

            self.students.append(Student(
                student_id=str(r["student_id"]),
                last_name=str(r["last_name"]),
                first_name=str(r["first_name"]),
                program=str(r["program"]),
                choices=choices,
                accommodation_sites=acc
            ))

        for _, r in sites_df.iterrows():
            s_name = str(r["site_name"]).strip()
            is_cardiac = bool(r["is_cardiac"])
            is_general = bool(r["is_general"])
            is_dual = bool(r["is_dual"])

            if is_cardiac: self.cardiac_names.add(s_name)
            if is_general: self.general_names.add(s_name)
            if is_dual: self.dual_site_names.add(s_name)

            self.sites.append(Site(
                site_id=r["site_id"],
                name=s_name,
                is_general=is_general,
                is_cardiac=is_cardiac,
                is_dual=is_dual,
                high_acuity=bool(r["high_acuity"]),
                out_of_town=bool(r["out_of_town"]),
                only_may_to_dec=bool(r["only_may_to_dec"]),
                capacity=int(r["capacity"])
            ))

        for _, r in clinics_df.iterrows():
            c_name = str(r["clinic_name"]).strip()
            self.clinic_names.add(c_name)
            self.clinics.append(Clinic(
                clinic_id=r["clinic_id"],
                name=c_name,
                days_per_week=int(r["days_per_week"]),
                out_of_town=bool(r["out_of_town"]),
                only_may_to_dec=bool(r["only_may_to_dec"])
            ))

    # ==================== SCHEDULING ====================

    def schedule_all(self):
        """
        Runs the complete scheduling process for all blocks.

        1. Resets previous state.
        2. Iterates through BLOCKS sequentially.
        3. Applies post-processing color overrides.
        4. Validates the schedule.
        5. Prints stats and returns the DataFrame.

        Returns:
            pd.DataFrame: The generated schedule suitable for Excel export.
        """
        self.reset()
        for block in BLOCKS:
            print(f"Scheduling Block {block}...")
            self.schedule_block(block)

        self.apply_color_overrides()
        self.validate()
        self.print_choice_stats()
        return self.generate_output()

    def get_match_count(self, s):
        """
        Calculates how many unique choice matches a student has received so far.

        Args:
            s (Student): The student to check.

        Returns:
            int: Number of placements that were in the student's top 5 choices.
        """
        assigned_names_this_block = set()
        for b_key, p_val in s.placements.items():
            clean = p_val.replace("*", "").split("(")[0].strip()
            assigned_names_this_block.add(clean)

        count = 0
        for choice in s.choices:
            if choice in assigned_names_this_block:
                count += 1
        return count

    def schedule_block(self, block):
        """
        Orchestrates the placement logic for a specific time block.

        Strategy:
        1. Identification: Filters active students for the block.
        2. Guaranteed Choice Phase (Priority 1): Prioritizes students with 0 matches
           so far to ensure equity.
        3. Regular Phase (Priority 2): Sorts remaining students based on program requirements
           (e.g., Dual students nearing cardiac limits, 2B/Finisher priority) and attempts
           placements.
        4. Fallback: If standard placement fails, attempts 'Emergency Place'.

        Args:
            block (str): The current block identifier (e.g., "2A", "4B").
        """
        active = [s for s in self.students if self.is_active(s, block)]

        # --- 1. GUARANTEED CHOICE PHASE ---
        zero_match_students = [
            s for s in active
            if self.get_match_count(s) == 0 and block not in s.placements and len(s.choices) > 0
        ]

        def zero_match_sort_key(s):
            """Sorting key for zero-match prioritization buckets."""
            bucket = 0
            if block in ["4A", "4B"]:
                if s.program == "Cardiac":
                    bucket += 1
                elif s.program == "Dual" and s.cardiac_weeks_completed < 12:
                    bucket += 2
                else:
                    bucket += 4
            else:
                if s.program == "Dual":
                    if 0 < s.cardiac_weeks_completed < 12:
                        bucket += 5
                    else:
                        bucket += 10
                else:
                    bucket += 50
            return bucket, random.random()

        random.shuffle(zero_match_students)
        zero_match_students.sort(key=zero_match_sort_key)

        for s in zero_match_students:
            if block in s.placements: continue

            # Ban clinics for Duals in 2A-3B (Worst case handling in place_student)
            placed = self.place_student(s, block, force_no_repeat=True, is_zero_match_priority=True, only_choices=True)
            if not placed:
                self.place_student(s, block, force_no_repeat=False, is_zero_match_priority=True,
                                   only_choices=True)

        # --- 2. REGULAR SCHEDULING ---
        remaining_active = [s for s in active if block not in s.placements]

        idx = BLOCKS.index(block)
        prev_block = BLOCKS[idx - 1] if idx > 0 else None

        def regular_sort_key(s):
            """Sorting key for regular placement based on program necessity and urgency."""
            must_continue = False
            if s.program == "Dual" and block in ["2B", "3B", "4B"] and s.is_chain_starter:
                must_continue = True

            bucket = 0 if must_continue else 100

            if block in ["4A", "4B"]:
                if s.program == "Cardiac":
                    bucket += 1
                elif s.program == "Dual" and s.cardiac_weeks_completed < 12:
                    bucket += 2
                else:
                    bucket += 4
            else:
                if s.program == "Dual":
                    # 2B Priority
                    if block == "2B" and s.cardiac_weeks_completed == 0:
                        bucket += 5
                    # Finisher Priority
                    elif 0 < s.cardiac_weeks_completed < 12:
                        bucket += 5
                    else:
                        bucket += 10
                else:
                    bucket += 50

            matches = self.get_match_count(s)
            return bucket, matches, random.random()

        random.shuffle(remaining_active)
        remaining_active.sort(key=regular_sort_key)

        # 2a. Clinic attempts for General (High Priority)
        for s in remaining_active:
            if s.program == "General" and block not in s.placements:
                if block == "3B" and s.clinic_blocks >= 3: continue
                if s.clinic_blocks < 3:
                    prob = 0.5
                    if block in ["2A", "2B"]:
                        prob = 0.90
                    elif block in ["3A", "3B"] and s.clinic_blocks == 0:
                        prob = 0.95
                    if random.random() < prob:
                        self.try_clinic(s, block, force_no_repeat=True)

        # 2b. Strict Phase
        for s in remaining_active:
            if block in s.placements: continue
            placed = self.place_student(s, block, force_no_repeat=True)
            if not placed and s.program == "General" and s.clinic_blocks < 3:
                self.try_clinic(s, block, force_no_repeat=True)

        # 2c. Fallback Phase
        for s in remaining_active:
            if block in s.placements: continue
            if s.program == "General":
                if block == "3B" and s.clinic_blocks >= 3: continue
                if s.clinic_blocks < 3 and random.random() < 0.5:
                    self.try_clinic(s, block, force_no_repeat=False)

        for s in remaining_active:
            if block in s.placements: continue
            placed = self.place_student(s, block, force_no_repeat=False)
            if not placed and s.program == "General" and s.clinic_blocks < 3:
                placed = self.try_clinic(s, block, force_no_repeat=False)

            if not placed:
                self.emergency_place(s, block)

    # ==================== COLOR OVERRIDES ====================

    def apply_color_overrides(self):
        """
        Post-processing step to update color intent based on specific rules.

        Rules:
        1. Orange: Applied to A/B blocks (e.g., 2A/2B) if the student is at the
           same Dual-capable site for both, indicating a Dual Chain.
        2. Purple: Applied if the site name contains "Women" or "BCWH".
        """
        for s in self.students:
            for a_block, b_block in [("2A", "2B"), ("3A", "3B"), ("4A", "4B")]:
                if a_block in s.placements and b_block in s.placements:
                    p1 = s.placements[a_block].replace("*", "").split("(")[0].strip()
                    p2 = s.placements[b_block].replace("*", "").split("(")[0].strip()
                    if p1 == p2 and p1 in self.dual_site_names:
                        self.placement_intent[(s.student_id, a_block)] = "Orange"
                        self.placement_intent[(s.student_id, b_block)] = "Orange"

            for block, placement in s.placements.items():
                p_clean = placement.replace("*", "").split("(")[0].strip()
                if "Women" in p_clean or "BCWH" in p_clean:
                    self.placement_intent[(s.student_id, block)] = "Purple"

    # ==================== PLACEMENT LOGIC ====================

    def place_student(self, s, block, force_no_repeat=False, is_zero_match_priority=False, only_choices=False,
                      ban_clinics=False):
        """
        Determines the placement strategy for a student based on their program and history.

        Args:
            s (Student): The student.
            block (str): The current block.
            force_no_repeat (bool): If True, disallows placing student at a site they've already visited.
            is_zero_match_priority (bool): If True, adjusts scoring to heavily favor student choices.
            only_choices (bool): If True, only considers sites listed in the student's choices.
            ban_clinics (bool): If True, prevents clinic assignment (used for certain Dual logic).

        Returns:
            bool: True if placement was successful, False otherwise.
        """
        if s.program == "Cardiac":
            return self.try_site(s, block, cardiac=True, force_no_repeat=force_no_repeat,
                                 is_zero_match_priority=is_zero_match_priority, only_choices=only_choices)

        if s.program == "Dual":
            # --- 1. CONTINUITY ---
            if block in ["2B", "3B", "4B"] and s.is_chain_starter:
                prev_block = BLOCKS[BLOCKS.index(block) - 1]
                prev_site = s.placements.get(prev_block, "")
                if "(" in prev_site: prev_site = prev_site.split("(")[0].strip()
                prev_site = prev_site.replace("*", "").strip()

                target_site = next((x for x in self.sites if x.name == prev_site), None)
                if target_site:
                    if self.site_name_usage[block][target_site.name] < target_site.capacity:
                        slot_idx = self.find_slot(target_site, block, 5)
                        if slot_idx is not None:
                            self.site_slots[block][target_site.site_id][slot_idx] -= 5
                            self.site_name_usage[block][target_site.name] += 1
                            self.assign_site(s, block, target_site, is_cardiac_intent=False)
                            s.is_chain_starter = False
                            return True

            # --- 2. 4A/4B LOGIC ---
            if block in ["4A", "4B"]:
                if s.cardiac_weeks_completed >= 12:
                    if self.try_clinic(s, block, force_no_repeat=force_no_repeat,
                                       is_zero_match_priority=is_zero_match_priority,
                                       only_choices=only_choices): return True
                    if self.try_site(s, block, general=True, force_no_repeat=force_no_repeat,
                                     is_zero_match_priority=is_zero_match_priority,
                                     only_choices=only_choices): return True
                    return False
                else:
                    if self.try_site(s, block, cardiac=True, force_no_repeat=force_no_repeat,
                                     is_zero_match_priority=is_zero_match_priority,
                                     only_choices=only_choices): return True
                    return self.try_clinic(s, block, force_no_repeat=force_no_repeat,
                                           is_zero_match_priority=is_zero_match_priority, only_choices=only_choices)

            # --- 3. 2A-3B LOGIC ---
            if s.cardiac_weeks_completed < 13:
                if self.is_same_number_cardiac_conflict(s, block):
                    return self.try_site(s, block, general=True, force_no_repeat=force_no_repeat,
                                         is_zero_match_priority=is_zero_match_priority, only_choices=only_choices)

                # STRANDED STUDENT: Force Cardiac
                if s.cardiac_weeks_completed > 0:
                    if self.try_site(s, block, cardiac=True, force_no_repeat=force_no_repeat,
                                     allow_new_dual_start=True, is_zero_match_priority=is_zero_match_priority,
                                     only_choices=only_choices): return True
                    return False

                # 0 Weeks: Load Balancer
                can_start_dual = (block in ["2A", "3A"])
                prefer_general = (random.random() < 0.5) if block in ["2A", "3A"] else False

                if prefer_general:
                    if self.try_site(s, block, general=True, force_no_repeat=force_no_repeat,
                                     is_zero_match_priority=is_zero_match_priority,
                                     only_choices=only_choices): return True
                    if self.try_site(s, block, cardiac=True, force_no_repeat=force_no_repeat,
                                     allow_new_dual_start=can_start_dual, is_zero_match_priority=is_zero_match_priority,
                                     only_choices=only_choices): return True
                else:
                    if self.try_site(s, block, cardiac=True, force_no_repeat=force_no_repeat,
                                     allow_new_dual_start=can_start_dual, is_zero_match_priority=is_zero_match_priority,
                                     only_choices=only_choices): return True
                    if self.try_site(s, block, general=True, force_no_repeat=force_no_repeat,
                                     is_zero_match_priority=is_zero_match_priority,
                                     only_choices=only_choices): return True

                # Worst Case: Clinic
                if not ban_clinics:
                    return self.try_clinic(s, block, force_no_repeat=force_no_repeat,
                                           is_zero_match_priority=is_zero_match_priority, only_choices=only_choices)
                return False

            # Done Cardiac, prefer General
            return self.try_site(s, block, general=True, force_no_repeat=force_no_repeat,
                                 is_zero_match_priority=is_zero_match_priority, only_choices=only_choices)

        if s.program == "General":
            # GENERALS = CLINICS ONLY (unless emergency)
            return self.try_clinic(s, block, force_no_repeat=force_no_repeat,
                                   is_zero_match_priority=is_zero_match_priority, only_choices=only_choices)
        return False

    def try_site(self, s, block, cardiac=False, general=False, force_no_repeat=False, allow_new_dual_start=True,
                 is_zero_match_priority=False, only_choices=False):
        """
        Attempts to assign a student to a hospital site.

        Logic includes:
        - Filtering based on site type (Cardiac/General).
        - Checking block availability (May-Dec vs Jan-Apr).
        - Validating OB exposure requirements.
        - Scoring candidates based on student choices, mandatory fills (BCWH),
          and out-of-town accommodation.

        Args:
            s (Student): Student to place.
            block (str): Block ID.
            cardiac (bool): Filter for cardiac sites.
            general (bool): Filter for general sites.
            force_no_repeat (bool): Prevent repeat visits.
            allow_new_dual_start (bool): Allows a Dual student to start a new chain at a Dual site.
            is_zero_match_priority (bool): Uses simplified scoring for students needing choices.
            only_choices (bool): Restrict candidates to student's choice list.

        Returns:
            bool: True if assigned, False otherwise.
        """
        candidates = []
        if cardiac: candidates.extend([x for x in self.sites if x.is_cardiac])
        if general: candidates.extend([x for x in self.sites if x.is_general])

        unique_map = {x.site_id: x for x in candidates}
        candidates = list(unique_map.values())

        # REMOVE SITES UNAVAILABLE IN SPRING
        if block in ["2A", "2B", "3A"]:
            candidates = [x for x in candidates if not x.only_may_to_dec]

        last_site_name = ""
        idx = BLOCKS.index(block)
        if idx > 0:
            prev_block = BLOCKS[idx - 1]
            last_site_name = s.placements.get(prev_block, "")
            if "(" in last_site_name: last_site_name = last_site_name.split("(")[0].strip()
            last_site_name = last_site_name.replace("*", "").strip()

        candidates = [x for x in candidates if self.allow_site(s, x)]

        filtered = []
        for x in candidates:
            is_continuity = (x.name == last_site_name)
            if x.is_dual and s.program == "Dual" and not is_continuity and not allow_new_dual_start:
                continue

            if only_choices and x.name not in s.choices:
                continue

            # Math Guard for Cardiac
            if s.program == "Dual" and cardiac:
                is_likely_cardiac_intent = False
                if x.is_cardiac:
                    if not x.is_general:
                        is_likely_cardiac_intent = True
                    elif cardiac:
                        is_likely_cardiac_intent = True

                if is_likely_cardiac_intent:
                    potential_total = s.cardiac_weeks_completed + BLOCK_WEEKS[block]
                    if potential_total > 13:
                        continue

            if force_no_repeat:
                if x.name in s.assigned_sites: continue
                filtered.append(x)
            else:
                filtered.append(x)

        candidates = filtered

        current_matches = self.get_match_count(s)
        ignore_choices_for_score = (current_matches >= 2) and not is_zero_match_priority

        def site_score(site):
            """Calculates a suitability score (lower is better, heavily modified by weights)."""
            score = 0

            if not is_zero_match_priority:
                score += (current_matches * 10000)

            # --- BC WOMEN'S MANDATORY FILL ---
            is_bcwh = "BCWH" in site.name
            if is_bcwh:
                current_fill = self.site_name_usage[block][site.name]

                # 1. FORCE FILL IF < 2 (Voluntold)
                if current_fill < 2:
                    score -= 10000000
                    # 2. Priority if chosen
                elif site.name in s.choices:
                    score -= 1000000
                # 3. Prefer Duals over Generals if not chosen
                elif s.program == "Dual":
                    score -= 5000
                else:
                    score += 50000

            if site.name in s.assigned_sites: score += 100000

            if s.program == "Dual" and site.is_dual:
                if site.name != last_site_name:
                    is_top_choice = False
                    try:
                        s.choices.index(site.name)
                        is_top_choice = True
                    except ValueError:
                        pass

                    if is_top_choice and not ignore_choices_for_score and self.dual_chains_assigned < self.MAX_DUAL_CHAINS:
                        score -= 500
                    else:
                        score += 1000

            if ignore_choices_for_score:
                score += 500
            else:
                try:
                    score += s.choices.index(site.name)
                except ValueError:
                    score += 500

            if site.out_of_town and site.name not in s.accommodation_sites: score += 500
            if not cardiac and site.is_cardiac: score += 1000

            return score, random.random()

        candidates.sort(key=site_score)

        for site in candidates:
            current_usage = self.site_name_usage[block][site.name]
            if current_usage >= site.capacity: continue

            slot_idx = self.find_slot(site, block, 5)
            if slot_idx is not None:
                self.site_slots[block][site.site_id][slot_idx] -= 5
                self.site_name_usage[block][site.name] += 1

                if s.program == "Dual" and site.is_dual and block in ["2A", "3A", "4A"]:
                    if self.dual_chains_assigned < self.MAX_DUAL_CHAINS:
                        s.is_chain_starter = True
                        self.dual_chains_assigned += 1

                is_actually_cardiac = False
                if site.is_cardiac and not site.is_general:
                    is_actually_cardiac = True
                elif site.is_cardiac and site.is_general:
                    if cardiac and not general:
                        is_actually_cardiac = True
                    elif general and not cardiac:
                        is_actually_cardiac = False
                    else:
                        if s.program == "Cardiac":
                            is_actually_cardiac = True
                        elif s.program == "Dual" and s.cardiac_weeks_completed < 13:
                            is_actually_cardiac = True
                        else:
                            is_actually_cardiac = False

                self.assign_site(s, block, site, is_actually_cardiac)
                return True
        return False

    def try_clinic(self, s, block, force_no_repeat=False, is_zero_match_priority=False, only_choices=False,
                   ban_clinics=False):
        """
        Attempts to assign a student to a clinic.

        Args:
            s (Student): Student to place.
            block (str): Block ID.
            force_no_repeat (bool): Prevent repeat visits.
            is_zero_match_priority (bool): High priority mode for students with no matches.
            only_choices (bool): Restrict to student's choices.
            ban_clinics (bool): Immediate exit flag.

        Returns:
            bool: True if assigned, False otherwise.
        """
        if ban_clinics: return False

        current_matches = self.get_match_count(s)
        ignore_choices_for_score = (current_matches >= 2) and not is_zero_match_priority

        def clinic_score(c):
            score = 0
            if not is_zero_match_priority:
                score += (current_matches * 10000)

            if c.name in s.assigned_sites: score += 100000

            if ignore_choices_for_score:
                score += 500
            else:
                try:
                    score += s.choices.index(c.name)
                except ValueError:
                    score += 500

            if c.out_of_town and c.name not in s.accommodation_sites: score += 500
            return score, random.random()

        candidates = self.clinics
        # REMOVE CLINICS UNAVAILABLE IN SPRING
        if block in ["2A", "2B", "3A"]:
            candidates = [c for c in candidates if not c.only_may_to_dec]

        if only_choices:
            candidates = [c for c in candidates if c.name in s.choices]

        if force_no_repeat:
            candidates = [c for c in candidates if c.name not in s.assigned_sites]
        sorted_clinics = sorted(candidates, key=clinic_score)

        for c in sorted_clinics:
            if self.clinic_days_used[block][c.clinic_id] > 0: continue
            self.assign_clinic(s, block, c)
            return True
        return False

    def emergency_place(self, s, block):
        """
        Last resort placement logic when all standard constraints fail.
        Relaxes program-specific constraints to find *any* slot.
        Assigns "UNASSIGNED" if truly no slot is found.
        """
        if s.program == "General":
            if self.try_clinic(s, block, force_no_repeat=False):
                self.placement_intent[(s.student_id, block)] = "Green"
                return
            if self.try_site(s, block, cardiac=True, general=True, force_no_repeat=False):
                self.placement_intent[(s.student_id, block)] = "Blue"
                return

        if s.program == "Dual":
            if s.cardiac_weeks_completed >= 12:
                if self.try_site(s, block, general=True, force_no_repeat=False):
                    self.placement_intent[(s.student_id, block)] = "Blue"
                    return
                if self.try_clinic(s, block, force_no_repeat=False):
                    self.placement_intent[(s.student_id, block)] = "Green"
                    return
            else:
                if self.try_site(s, block, cardiac=True, force_no_repeat=False):
                    self.placement_intent[(s.student_id, block)] = "Red"
                    return
                if self.try_site(s, block, general=True, force_no_repeat=False):
                    self.placement_intent[(s.student_id, block)] = "Blue"
                    return
                if self.try_clinic(s, block, force_no_repeat=False):
                    self.placement_intent[(s.student_id, block)] = "Green"
                    return

        s.placements[block] = "UNASSIGNED"
        self.placement_intent[(s.student_id, block)] = "Warn"

    # ==================== HELPERS ====================

    def is_same_number_cardiac_conflict(self, s, block):
        """
        Checks if the previous block was cardiac and assigning another cardiac
        block with the same number (e.g., 2A -> 2B) might cause issues.
        (Logic specific to split blocks).
        """
        idx = BLOCKS.index(block)
        if idx == 0: return False
        prev_block = BLOCKS[idx - 1]

        prev_intent = self.placement_intent.get((s.student_id, prev_block))

        if prev_intent == "Red":
            if block[0] == prev_block[0]:
                return True
        return False

    def find_slot(self, site, block, days_needed):
        """
        Finds an index in the site's capacity tracking list that has enough days remaining.

        Args:
            site (Site): The site to check.
            block (str): The block to check.
            days_needed (int): Number of days needed (usually 5).

        Returns:
            int or None: The index of the available slot, or None if full.
        """
        slots = self.site_slots[block][site.site_id]
        for i, remaining in enumerate(slots):
            if remaining >= days_needed: return i
        return None

    def assign_site(self, s, block, site, is_cardiac_intent):
        """
        Finalizes a student's assignment to a site. Updates student history and flags.
        """
        marker = "*" if site.name in s.choices else ""
        s.placements[block] = f"{site.name}{marker}"
        s.assigned_sites.append(site.name)

        if is_cardiac_intent:
            self.placement_intent[(s.student_id, block)] = "Red"
            s.cardiac_weeks_completed += BLOCK_WEEKS[block]
        else:
            self.placement_intent[(s.student_id, block)] = "Blue"

        if not is_cardiac_intent and site.name != "Vancouver General":
            s.has_had_ob_exposure = True

    def assign_clinic(self, s, block, clinic):
        """
        Finalizes a student's assignment to a clinic. Updates student history and usage counters.
        """
        self.clinic_days_used[block][clinic.clinic_id] += clinic.days_per_week
        marker = "*" if clinic.name in s.choices else ""
        s.placements[block] = f"{clinic.name}{marker} ({clinic.days_per_week}d)"
        s.assigned_sites.append(clinic.name)
        s.clinic_blocks += 1
        s.has_had_ob_exposure = True
        self.placement_intent[(s.student_id, block)] = "Green"

    def allow_site(self, s, site):
        """
        Checks specific hard constraints for a site.
        Example: Vancouver General requires prior OB exposure.
        """
        if site.name == "Vancouver General" and not s.has_had_ob_exposure: return False
        return True

    def is_active(self, s, block):
        """
        Determines if a student is active (requires placement) in a given block
        based on their program stream.
        """
        if s.program == "General" and block not in ["2A", "2B", "3A", "3B"]: return False
        if s.program == "Dual" and block == "Summer": return False
        if s.program == "Cardiac" and block in ["2A", "2B", "3A", "3B"]: return False
        return True

    def reset(self):
        """Clears all scheduling state to allow for a fresh run."""
        self.placement_intent.clear()
        self.dual_chains_assigned = 0
        for s in self.students:
            s.cardiac_weeks_completed = 0
            s.clinic_blocks = 0
            s.is_chain_starter = False
            s.has_had_ob_exposure = False
            s.placements.clear()
            s.assigned_sites.clear()
        for b in BLOCKS:
            self.clinic_days_used[b].clear()
            self.site_name_usage[b].clear()
            self.site_slots[b].clear()
            for site in self.sites:
                self.site_slots[b][site.site_id] = [5] * site.capacity

    def print_choice_stats(self):
        """Prints a summary of how well students' choices were met to the console."""
        print("\n" + "=" * 40)
        print("       STUDENT CHOICE SATISFACTION       ")
        print("=" * 40)

        total_students = len(self.students)
        total_matches = 0
        students_with_0 = 0

        print(f"{'Name':<25} | {'Matches':<8} | {'Matched Choices'}")
        print("-" * 65)

        for s in self.students:
            matches = self.get_match_count(s)
            matched_list = []
            for p in s.placements.values():
                clean = p.replace("*", "").split("(")[0].strip()
                if clean in s.choices: matched_list.append(clean)

            total_matches += matches
            if matches == 0:
                students_with_0 += 1

            print(f"{s.first_name} {s.last_name:<15} | {matches}/5     | {', '.join(set(matched_list))}")

        avg = total_matches / total_students if total_students > 0 else 0
        print("-" * 65)
        print(f"Total Students: {total_students}")
        print(f"Average Choices Met: {avg:.2f} / 5.0")
        print(f"Students with 0 matches: {students_with_0}")
        print("=" * 40 + "\n")

    def validate(self):
        """
        Validates the final schedule for logic errors, such as unassigned slots
        or Dual students failing to meet cardiac week minimums.
        """
        print("\n--- VALIDATION ---")
        unassigned_count = 0
        for s in self.students:
            if s.program == "Dual":
                if s.cardiac_weeks_completed < 12:
                    print(f"⚠️ {s.first_name} (Dual) only has {s.cardiac_weeks_completed} cardiac weeks.")

            site_counts = pd.Series(s.assigned_sites).value_counts()
            repeats = site_counts[site_counts > 1]
            if not repeats.empty:
                pass

            for b in BLOCKS:
                if self.is_active(s, b):
                    if b not in s.placements or "UNASSIGNED" in s.placements[b]:
                        print(f"⚠️ {s.first_name} unassigned in {b}")
                        unassigned_count += 1

        if unassigned_count == 0:
            print("✅ All active slots filled.")
        else:
            print(f"❌ {unassigned_count} unassigned slots found.")

    def generate_output(self):
        """
        Constructs a Pandas DataFrame representing the final schedule.
        Also appends a list of unused sites/capacities at the bottom.
        """
        global row
        order = {"General": 0, "Dual": 1, "Cardiac": 2}
        rows = []
        for s in sorted(self.students, key=lambda x: order.get(x.program, 3)):
            row = {
                "student_id": s.student_id,
                "last_name": s.last_name,
                "first_name": s.first_name,
                "program": s.program
            }
            for b in BLOCKS:
                row[b] = s.placements.get(b, "")
            rows.append(row)

        # Add Unused Sites List
        empty_row = {k: "" for k in row.keys()}
        rows.append(empty_row)
        rows.append({**empty_row, "student_id": "UNUSED SITES"})

        max_unused = 0
        unused_map = defaultdict(list)

        for b in BLOCKS:
            for site in self.sites:
                if b in ["2A", "2B", "3A"] and site.only_may_to_dec: continue
                used = self.site_name_usage[b][site.name]
                remaining = site.capacity - used
                if remaining > 0:
                    unused_map[b].append(f"{site.name} ({remaining})")

            for c in self.clinics:
                if b in ["2A", "2B", "3A"] and c.only_may_to_dec: continue
                if self.clinic_days_used[b][c.clinic_id] == 0:
                    unused_map[b].append(f"{c.name} (Open)")

            if len(unused_map[b]) > max_unused:
                max_unused = len(unused_map[b])

        for i in range(max_unused):
            u_row = {**empty_row}
            for b in BLOCKS:
                if i < len(unused_map[b]):
                    u_row[b] = unused_map[b][i]
            rows.append(u_row)

        return pd.DataFrame(rows)


# ==================== EXCEL COLORING ====================

def apply_formatting(filename, scheduler):
    """
    Applies conditional formatting to the generated Excel file using OpenPyXL.

    Color Scheme:
    - Red: Cardiac placement.
    - Blue: General placement.
    - Green: Clinic placement.
    - Orange: Dual placement chain (same site for A & B blocks).
    - Purple: BCWH placement.
    - Warn (Yellow): Issues or Unassigned slots.

    Args:
        filename (str): Path to the Excel file to format.
        scheduler (Scheduler): The scheduler instance containing placement intents.
    """
    print(f"Applying color coding to {filename}...")
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except Exception as e:
        print(f"Error opening workbook for styling: {e}")
        return

    # Styles
    fill_cardiac = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    fill_clinic = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    fill_general = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Blue
    fill_dual = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # Peachy Orange
    fill_bcwh = PatternFill(start_color="CC99FF", end_color="CC99FF", fill_type="solid")  # Light Lavender
    fill_warn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header Mapping
    header_row = ws[1]
    block_col_indices = {}
    for cell in header_row:
        cell.border = thin_border
        if cell.value in BLOCKS:
            block_col_indices[cell.col_idx] = cell.value

    # Row Iteration
    for row in ws.iter_rows(min_row=2):
        cell_val_A = str(row[0].value) if row[0].value else ""
        if cell_val_A == "UNUSED SITES": break

        student_id = cell_val_A

        for cell in row:
            cell.border = thin_border

            if cell.col_idx in block_col_indices:
                block = block_col_indices[cell.col_idx]
                intent = scheduler.placement_intent.get((student_id, block))

                if intent == "Red":
                    cell.fill = fill_cardiac
                elif intent == "Blue":
                    cell.fill = fill_general
                elif intent == "Green":
                    cell.fill = fill_clinic
                elif intent == "Orange":
                    cell.fill = fill_dual
                elif intent == "Purple":
                    cell.fill = fill_bcwh
                elif intent == "Warn":
                    cell.fill = fill_warn

                if "UNASSIGNED" in str(cell.value):
                    cell.fill = fill_warn

    wb.save(filename)
    print("✅ Color coding and borders applied.")


# ==================== CONSTANTS ====================

BLOCKS = ["2A", "2B", "3A", "3B", "Summer", "4A", "4B"]
BLOCK_WEEKS = {"2A": 7, "2B": 6, "3A": 6, "3B": 7, "Summer": 7, "4A": 7, "4B": 6}


# ==================== MAIN ====================

def main():
    """
    Main entry point for the script.
    Loads data, runs the scheduler, saves the Excel file, and applies formatting.
    """
    scheduler = Scheduler("Student_Site_Data.xlsx")
    scheduler.load_data()
    if not scheduler.students:
        print("No students loaded. Check Excel file.")
        return

    df = scheduler.schedule_all()
    output_file = "Generated_Schedule.xlsx"
    df.to_excel(output_file, index=False)

    apply_formatting(output_file, scheduler)
    print(f"Schedule saved to {output_file}")


if __name__ == "__main__":
    main()